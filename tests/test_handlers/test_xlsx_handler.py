"""Tests for XlsxHandler — XLSX text extraction."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from core.result import Failure, Success
from handlers.xlsx_handler import XlsxHandler


@pytest.fixture
def xlsx_path(tmp_path: Path) -> Path:
    """Single-sheet workbook with header + two data rows."""
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Revenue"
    ws.append(["Date", "Product", "Amount"])
    ws.append(["2026-01-01", "Widget A", 1200])
    ws.append(["2026-01-02", "Widget B", 800])
    wb.save(str(path))
    return path


@pytest.fixture
def multi_sheet_path(tmp_path: Path) -> Path:
    """Workbook with two named sheets."""
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Q1"
    ws1.append(["Month", "Sales"])
    ws1.append(["Jan", 100])
    ws2 = wb.create_sheet("Q2")
    ws2.append(["Month", "Sales"])
    ws2.append(["Apr", 200])
    wb.save(str(path))
    return path


@pytest.fixture
def empty_sheet_path(tmp_path: Path) -> Path:
    """Workbook: one sheet with data, one empty sheet."""
    path = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Data"
    ws1.append(["Name", "Value"])
    ws1.append(["A", 1])
    wb.create_sheet("Empty")
    wb.save(str(path))
    return path


@pytest.fixture
def all_empty_path(tmp_path: Path) -> Path:
    """Workbook where all sheets have no rows."""
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    wb.save(str(path))
    return path


@pytest.fixture
def formula_path(tmp_path: Path) -> Path:
    """Workbook with a cell that has a cached value (data_only reads value)."""
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Calc"
    ws.append(["A", "B", "Sum"])
    ws["C2"] = 42
    wb.save(str(path))
    return path


class TestXlsxHandlerCanHandle:
    def test_lowercase_xlsx(self) -> None:
        assert XlsxHandler().can_handle(Path("file.xlsx")) is True

    def test_uppercase_xlsx(self) -> None:
        assert XlsxHandler().can_handle(Path("file.XLSX")) is True

    def test_xls_not_handled(self) -> None:
        assert XlsxHandler().can_handle(Path("file.xls")) is False

    def test_csv_not_handled(self) -> None:
        assert XlsxHandler().can_handle(Path("file.csv")) is False


class TestXlsxHandlerExtract:
    def test_single_sheet_has_sheet_header(self, xlsx_path: Path) -> None:
        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Success)
        assert '[Sheet: "Revenue"]' in result.value.text

    def test_single_sheet_has_column_headers(self, xlsx_path: Path) -> None:
        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Success)
        assert "Date" in result.value.text
        assert "Product" in result.value.text
        assert "Amount" in result.value.text

    def test_single_sheet_has_data_rows(self, xlsx_path: Path) -> None:
        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Success)
        assert "Widget A" in result.value.text
        assert "1200" in result.value.text

    def test_single_sheet_is_md_false(self, xlsx_path: Path) -> None:
        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Success)
        assert result.value.is_md is False

    def test_single_sheet_source_path(self, xlsx_path: Path) -> None:
        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Success)
        assert result.value.source_path == xlsx_path

    def test_multi_sheet_both_headers_present(self, multi_sheet_path: Path) -> None:
        result = XlsxHandler().extract(multi_sheet_path)
        assert isinstance(result, Success)
        assert '[Sheet: "Q1"]' in result.value.text
        assert '[Sheet: "Q2"]' in result.value.text

    def test_multi_sheet_both_data_present(self, multi_sheet_path: Path) -> None:
        result = XlsxHandler().extract(multi_sheet_path)
        assert isinstance(result, Success)
        assert "100" in result.value.text
        assert "200" in result.value.text

    def test_empty_sheet_absent_from_text(self, empty_sheet_path: Path) -> None:
        result = XlsxHandler().extract(empty_sheet_path)
        assert isinstance(result, Success)
        assert '[Sheet: "Data"]' in result.value.text
        assert '[Sheet: "Empty"]' not in result.value.text

    def test_all_empty_workbook_returns_success_empty_text(
        self, all_empty_path: Path
    ) -> None:
        result = XlsxHandler().extract(all_empty_path)
        assert isinstance(result, Success)
        assert result.value.text == ""

    def test_nonexistent_path_returns_failure(self, tmp_path: Path) -> None:
        result = XlsxHandler().extract(tmp_path / "ghost.xlsx")
        assert isinstance(result, Failure)
        assert result.recoverable is False

    def test_file_too_large_returns_failure(
        self, xlsx_path: Path, monkeypatch
    ) -> None:
        import core.config as cfg_module
        from unittest.mock import MagicMock

        import handlers.xlsx_handler as xlsx_module
        from core.config import HandlersConfig

        fake = MagicMock()
        fake.main.handlers = HandlersConfig(max_file_size_bytes=1)
        monkeypatch.setattr(cfg_module, "_CONFIG", fake)

        # Guard rejects before any parse — load_workbook must not be reached.
        def _boom(*args, **kwargs):  # pragma: no cover - must not run
            raise AssertionError("load_workbook should not be called")

        monkeypatch.setattr(xlsx_module.openpyxl, "load_workbook", _boom)

        result = XlsxHandler().extract(xlsx_path)
        assert isinstance(result, Failure)
        assert result.recoverable is False
        assert "too large" in result.error
